"""
4910 SSA 데일리 리포트 자동화
- 매일 전일자 Airbridge 데이터를 Google Drive xlsx에 적재
"""
import io, json, time, zipfile, re, urllib.request, os
from datetime import datetime, timedelta, date
import pandas as pd
import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

# ── 설정 ──────────────────────────────────────────────────────────
AB_TOKEN  = os.environ["AIRBRIDGE_TOKEN"]
FOLDER_ID = os.environ["DRIVE_FOLDER_ID"]
GCP_KEY   = json.loads(os.environ["GCP_SERVICE_ACCOUNT_JSON"])

TARGET = (datetime.utcnow() + timedelta(hours=9) - timedelta(days=1)).strftime("%Y-%m-%d")
print(f"적재 대상 날짜: {TARGET}")

# ── Google Drive 연결 ──────────────────────────────────────────────
creds = service_account.Credentials.from_service_account_info(
    GCP_KEY, scopes=["https://www.googleapis.com/auth/drive"]
)
drive = build("drive", "v3", credentials=creds)

def list_drive_files():
    return drive.files().list(
        q=f"'{FOLDER_ID}' in parents and trashed=false",
        fields="files(id,name,modifiedTime)",
        orderBy="modifiedTime desc",
        supportsAllDrives=True, includeItemsFromAllDrives=True
    ).execute().get("files", [])

# ── 최신 파일 찾기 ─────────────────────────────────────────────────
files = list_drive_files()
src = files[0]
print(f"소스 파일: {src['name']}")

# 이미 오늘 날짜 파일이 있으면 스킵
month_day = TARGET[5:].replace("-", ".")  # "06.25"
if any(month_day in f["name"] for f in files):
    print(f"이미 {month_day} 파일 존재. 스킵.")
    exit(0)

# ── 소스 다운로드 ──────────────────────────────────────────────────
buf = io.BytesIO()
dl = MediaIoBaseDownload(buf, drive.files().get_media(fileId=src["id"], supportsAllDrives=True))
done = False
while not done: _, done = dl.next_chunk()
xlsx_bytes = buf.getvalue()
print(f"다운로드: {len(xlsx_bytes):,} bytes")

# ── Airbridge API 데이터 추출 ──────────────────────────────────────
HEADERS = {"Authorization": f"Bearer {AB_TOKEN}", "Content-Type": "application/json"}
BODY = {
    "groupBys": ["event_date","campaign","ad_group","ad_creative"],
    "metrics": ["cost_channel","impressions_channel","clicks_channel","web_opens",
                "app_web_sign_up","app_installs","app_first_installs",
                "app_custom_users_COMPLETE_ORDER_DOMESTIC","web_custom_users_COMPLETE_ORDER_DOMESTIC",
                "app_web_order_complete","app_web_revenue","web_order_complete_users","app_order_complete_users"],
    "filters": [
        {"field":"channel","filterType":"IN","values":["naver.searchad"]},
        {"field":"campaign","filterType":"IN","values":["구매 캠페인"]},
        {"field":"ad_group","filterType":"LIKE","values":["*네이버_검색광고"]}
    ],
    "sorts": [{"fieldName":"event_date","isAscending":False}],
    "option": {"timezone": None, "eventTimestampSource": "event_occurred_date"},
    "size": 500000, "isSummaryAvailable": True, "viewFormat": True, "skip": 0,
    "from": TARGET, "to": TARGET
}

with urllib.request.urlopen(urllib.request.Request(
    "https://api.airbridge.io/reports/api/v7/apps/4910/actuals/query",
    data=json.dumps(BODY).encode(), headers=HEADERS, method="POST"), timeout=15) as r:
    task_id = json.loads(r.read())["task"]["taskId"]
print(f"Airbridge 쿼리: {task_id}")

for _ in range(40):
    time.sleep(3)
    with urllib.request.urlopen(urllib.request.Request(
        f"https://api.airbridge.io/reports/api/v7/apps/4910/actuals/query/{task_id}",
        headers=HEADERS), timeout=10) as r:
        status = json.loads(r.read())["task"]["status"]
    if status == "SUCCESS":
        print("데이터 준비 완료")
        break
    print(f"  대기 중... ({status})")

with urllib.request.urlopen(urllib.request.Request(
    f"https://api.airbridge.io/reports/api/v7/apps/4910/actuals/export/{task_id}"
    f"?skip=0&size=500000&isSummaryAvailable=true&viewFormat=true",
    headers=HEADERS), timeout=10) as r:
    s3_url = json.loads(r.read())["url"]

csv_bytes = urllib.request.urlopen(s3_url, timeout=60).read()
df = pd.read_csv(io.BytesIO(csv_bytes), encoding="utf-8-sig")
df = df[df["Event Date"] == TARGET].reset_index(drop=True)
n = len(df)
print(f"추출 행수: {n}")
if n == 0:
    print("데이터 없음. 종료.")
    exit(0)

# ── Excel 적재 ────────────────────────────────────────────────────
print("Excel 로딩 중...")
wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), keep_vba=False)
ws = wb["RD"]
ws.sheet_state = "visible"

# 마지막 행 확인
last_row = 1
for r in range(ws.max_row, 1, -1):
    if ws.cell(r, 9).value is not None:
        last_row = r
        break
print(f"적재 시작 행: {last_row + 1}")

# I열 날짜 서식 복사
date_fmt = ws.cell(last_row, 9).number_format
fmt_cache = {col: ws.cell(last_row, col).number_format for col in range(9, 26)}

TARGET_DT = datetime.strptime(TARGET, "%Y-%m-%d")
CSV_COLS = [
    "Event Date","Campaign","Ad Group","Ad Creative",
    "Cost (Channel)","Impressions (Channel)","Clicks (Channel)",
    "Opens (Web)","Sign-up (App+Web)","Installs (App)",
    "First Installs (App)","COMPLETE_ORDER_DOMESTIC Users (App)",
    "COMPLETE_ORDER_DOMESTIC Users (Web)","Order Complete (App+Web)",
    "Revenue (App+Web)","Order Complete Users (Web)","Order Complete Users (App)",
]
STR_COLS = {"Campaign","Ad Group","Ad Creative"}

for i, row_data in df.iterrows():
    excel_row = last_row + 1 + i
    for col_idx, col_name in enumerate(CSV_COLS, start=9):
        cell = ws.cell(excel_row, col_idx)
        cell.number_format = fmt_cache[col_idx]
        if col_name == "Event Date":
            cell.value = TARGET_DT
        elif col_name in STR_COLS:
            v = row_data[col_name]
            cell.value = str(v) if pd.notna(v) else ""
        else:
            v = row_data[col_name]
            cell.value = float(v) if pd.notna(v) else 0.0

# A:H, Z 수식 복사
FORMULA_COLS = list(range(1, 9)) + [26]
formulas = {col: ws.cell(2, col).value for col in FORMULA_COLS}
G_TMPL = "=IF(COUNTIF('운영 중인 그룹'!$B$2:$B$191,K{r})>0,\"Y\",\"N\")"
H_TMPL = '=IFERROR(MID(L{r}, FIND("[", L{r})+1, FIND("]", L{r})-FIND("[", L{r})-1), "")'

first_new = last_row + 1
new_last  = last_row + n
for r in range(first_new, new_last + 1):
    for col in FORMULA_COLS:
        cell = ws.cell(r, col)
        if col == 7:
            cell.value = G_TMPL.format(r=r)
        elif col == 8:
            cell.value = H_TMPL.format(r=r)
        else:
            cell.value = formulas[col]

# 테이블 범위 확장
for tname in ws.tables:
    tbl = ws.tables[tname]
    col_letters = ''.join(c for c in tbl.ref.split(":")[1] if c.isalpha())
    tbl.ref = f"{tbl.ref.split(':')[0]}:{col_letters}{new_last}"
    print(f"테이블 '{tname}' → {tbl.ref}")

ws.sheet_state = "hidden"
print(f"적재 완료: {n}행")

# ── pivotCache refreshOnLoad=1 ─────────────────────────────────────
out_buf = io.BytesIO()
wb.save(out_buf)
out_bytes = out_buf.getvalue()

final_buf = io.BytesIO()
with zipfile.ZipFile(io.BytesIO(out_bytes), 'r') as zin, \
     zipfile.ZipFile(final_buf, 'w', zipfile.ZIP_DEFLATED) as zout:
    for item in zin.infolist():
        data = zin.read(item.filename)
        if 'pivotCacheDefinition' in item.filename:
            xml = data.decode('utf-8')
            xml = re.sub(r'refreshOnLoad="0"', 'refreshOnLoad="1"', xml)
            if 'refreshOnLoad' not in xml:
                xml = xml.replace('<pivotCacheDefinition ', '<pivotCacheDefinition refreshOnLoad="1" ', 1)
            data = xml.encode('utf-8')
        zout.writestr(item, data)
final_bytes = final_buf.getvalue()

# ── Drive 업로드 ───────────────────────────────────────────────────
# 파일명: 소스에서 날짜 부분 교체 (예: 6.24 → 6.25)
src_month_day = src["name"][src["name"].rfind("_")+1:].replace(".xlsx","")  # "6.24"
new_month_day = f"{int(TARGET[5:7])}.{int(TARGET[8:10])}"                   # "6.25"
new_name = src["name"].replace(src_month_day, new_month_day)
print(f"업로드: {new_name}")

media = MediaIoBaseUpload(io.BytesIO(final_bytes),
    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    resumable=True)
uploaded = drive.files().create(
    body={"name": new_name, "parents": [FOLDER_ID]},
    media_body=media, fields="id,name",
    supportsAllDrives=True
).execute()
print(f"완료: {uploaded['name']} ({uploaded['id']})")
